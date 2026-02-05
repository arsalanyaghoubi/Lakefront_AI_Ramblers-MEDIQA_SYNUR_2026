import argparse
import json
import csv
import subprocess
import tempfile
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

SCRIPT_DIR = Path(__file__).parent.resolve()
OUTPUTS_DIR = SCRIPT_DIR / "outputs"
PROJECT_ROOT = SCRIPT_DIR.parent
SCHEMA_PATH = PROJECT_ROOT / "Data" / "synur_schema.json"
TRAIN_PATH = PROJECT_ROOT / "Data" / "train.jsonl"
DEV_PATH = PROJECT_ROOT / "Data" / "dev.jsonl"


def load_schema():

    with open(SCHEMA_PATH) as f:
        schema = json.load(f)
    return {str(s['id']): s for s in schema}


def load_reference_data(data_split):

    data_path = DEV_PATH if data_split == "dev" else TRAIN_PATH
    if not data_path.exists():
        return {}
    
    gold_by_transcript = {}
    with open(data_path) as f:
        for line in f:
            record = json.loads(line)
            tid = str(record['id'])
            obs = record.get('observations', [])
            if isinstance(obs, str):
                obs = json.loads(obs)
            gold_by_transcript[tid] = obs
    return gold_by_transcript


def run_official_eval(pred_jsonl_path, ref_jsonl_path):
    """Run official evaluation script."""
    eval_script = PROJECT_ROOT / "mediqa_synur_eval_script.py"
    
    result = subprocess.run(
        ["python", str(eval_script), "-r", str(ref_jsonl_path), "-p", str(pred_jsonl_path)],
        capture_output=True,
        text=True
    )
    
    lines = result.stdout.strip().split("\n")
    try:
        idx = lines.index("Final Results:")
        precision = float(lines[idx + 1])
        recall = float(lines[idx + 2])
        f1 = float(lines[idx + 3])
        return {"precision": precision, "recall": recall, "f1": f1}
    except (ValueError, IndexError):
        return None


def evaluate_submission_file(submission_path, data_split):
    """Evaluate a submission file against gold labels."""
    ref_path = DEV_PATH if data_split == "dev" else TRAIN_PATH
    if not ref_path.exists():
        return None
    
    return run_official_eval(str(submission_path), str(ref_path))


def discover_experiments():

    experiments = []
    
    for summary_file in OUTPUTS_DIR.glob("**/summary_*.json"):
        exp = parse_summary_file(summary_file, legacy=True)
        if exp:
            experiments.append(exp)
    
    # Pattern 2: New structure - run_*/summary.json
    for summary_file in OUTPUTS_DIR.glob("**/run_*/summary.json"):
        exp = parse_summary_file(summary_file, legacy=False)
        if exp:
            experiments.append(exp)
    
    return experiments


def parse_summary_file(summary_file, legacy=True):
    """Parse a summary file and return experiment dict."""
    try:
        with open(summary_file) as f:
            content = f.read().strip()
            if not content:
                return None
            summary = json.loads(content)
        
        # Skip test runs with limit
        if summary.get("limit") is not None:
            return None
        
        # Skip temperature > 0 runs (used only for SC voting)
        temp = summary.get("temperature", 0)
        if temp > 0:
            return None
        
        timestamp = summary.get("timestamp", "")
        
        # Find corresponding files based on structure
        if legacy:
            pred_file = summary_file.parent / f"predictions_{timestamp}.json"
            submission_file = summary_file.parent / f"submission_{timestamp}.jsonl"
        else:
            # New structure: files are in same run_* folder without timestamps
            pred_file = summary_file.parent / "predictions.json"
            submission_file = summary_file.parent / "submission.jsonl"
        
        # Determine data_split from num_transcripts if unknown
        data_split = summary.get("data_split", "unknown")
        num_transcripts = summary.get("num_transcripts", 0)
        if data_split == "unknown":
            if num_transcripts == 101:
                data_split = "dev"
            elif num_transcripts == 122:
                data_split = "train"
        
        exp = {
            "timestamp": timestamp,
            "data_split": data_split,
            "model": summary.get("llm_model", summary.get("model_display_name", "unknown")),
            "model_path": summary.get("llm_model_path", ""),
            "embedding": summary.get("embedding_model", "").split("/")[-1],
            "hybrid": summary.get("hybrid_search", False),
            "few_shot": summary.get("few_shot", False),
            "num_examples": summary.get("num_few_shot_examples", 0),
            "segmentation": summary.get("segmentation", "simple"),
            "quantization": summary.get("quantization"),
            "num_transcripts": num_transcripts,
            "precision": summary.get("precision"),
            "recall": summary.get("recall"),
            "f1": summary.get("f1"),
            "source": "extraction",
            "submission_file": submission_file if submission_file.exists() else None,
        }
        
        exp["full_name"] = build_full_name(exp)
        return exp
        
    except Exception as e:
        print(f"Warning: Could not parse {summary_file}: {e}")
        return None


def discover_post_processing_results():
    """Discover voting and verification results."""
    results = []
    
    # Look for voted_* and verified_* submission files
    for submission_file in OUTPUTS_DIR.glob("**/voted_submission*.jsonl"):
        result = process_post_processing_file(submission_file, "voting")
        if result:
            results.append(result)
    
    # Match both verified_submission* and verified_voted_submission*
    for submission_file in OUTPUTS_DIR.glob("**/verified_*submission*.jsonl"):
        result = process_post_processing_file(submission_file, "verification")
        if result:
            results.append(result)
    
    return results


def process_post_processing_file(submission_file, process_type):
    """Process a post-processing result file."""
    try:
        # Count transcripts to determine data split
        num_transcripts = 0
        with open(submission_file) as f:
            for line in f:
                if line.strip():
                    num_transcripts += 1
        
        if num_transcripts == 101:
            data_split = "dev"
        elif num_transcripts == 122:
            data_split = "train"
        else:
            data_split = "unknown"
        
        # Evaluate
        metrics = evaluate_submission_file(submission_file, data_split)
        if not metrics:
            return None
        
        # Determine the base model from path
        path_parts = submission_file.parts
        model_folder = None
        for part in path_parts:
            if "Llama" in part or "gpt" in part.lower():
                model_folder = part
                break
        
        # Build name based on file name
        file_stem = submission_file.stem
        
        if process_type == "voting":
            if "sc123" in file_stem.lower():
                name = "SC Voting (2/3 majority)"
            else:
                name = "SC Voting"
        else:  # verification
            # Determine verifier from filename
            if "70b" in file_stem.lower():
                verifier = "Llama-70B"
            elif "segment" in file_stem.lower():
                verifier = "GPT-4o-mini (segment-level)"
            else:
                verifier = "GPT-4o-mini"
            
            if "voted" in file_stem.lower():
                name = f"SC Voting + Verification ({verifier})"
            else:
                name = f"Verification ({verifier})"
        
        # Determine base model
        if model_folder and "sft" in model_folder.lower():
            base_model = "Llama-70B-SFT"
        elif model_folder and "70b" in model_folder.lower():
            base_model = "Llama-70B"
        else:
            base_model = "Unknown"
        
        return {
            "timestamp": submission_file.stat().st_mtime,
            "data_split": data_split,
            "model": base_model,
            "embedding": "BGE",
            "hybrid": True,
            "few_shot": False,
            "num_examples": 0,
            "segmentation": "simple",
            "quantization": "4-bit",
            "num_transcripts": num_transcripts,
            "precision": metrics["precision"],
            "recall": metrics["recall"],
            "f1": metrics["f1"],
            "source": process_type,
            "full_name": f"{base_model} + {name}",
            "submission_file": submission_file,
        }
    except Exception as e:
        print(f"Warning: Could not process {submission_file}: {e}")
        return None


def build_full_name(exp):

    parts = []
    
    model = exp["model"]
    model_path = exp.get("model_path", "")
    
    # Determine model name
    if "gpt-4o-mini" in model.lower():
        parts.append("GPT-4o-mini")
    elif "gpt-4o" in model.lower():
        parts.append("GPT-4o")
    elif "sft" in model.lower() or "sft" in str(model_path).lower():
        if "70b" in model.lower() or "70b" in str(model_path).lower():
            parts.append("Llama-70B-SFT")
        else:
            parts.append("Llama-8B-SFT")
    elif "70b" in model.lower():
        parts.append("Llama-70B-Base")
    elif "8b" in model.lower() or "llama-3" in model.lower():
        parts.append("Llama-8B-Base")
    else:
        parts.append(model[:20] if len(model) > 20 else model)
    
    # Embedding
    emb = exp.get("embedding", "")
    if "bge" in emb.lower():
        parts.append("BGE")
    elif "minilm" in emb.lower():
        parts.append("MiniLM")
    else:
        parts.append("OpenAI")
    
    # Retrieval
    parts.append("Hybrid" if exp.get("hybrid") else "Dense")
    
    # Segmentation
    if exp.get("segmentation") == "llm":
        parts.append("LLM-seg")
    
    # Few-shot
    if exp.get("few_shot"):
        parts.append(f"{exp.get('num_examples', 0)}-shot")
    else:
        parts.append("0-shot")
    
    return " + ".join(parts)


def normalize_value(value):

    if value is None:
        return None
    if isinstance(value, str):
        return value.strip().lower()
    if isinstance(value, list):
        return sorted([normalize_value(v) for v in value])
    if isinstance(value, (int, float)):
        return float(value)
    return str(value).strip().lower()


def compute_error_analysis(exp, gold_by_transcript):
    """Compute error analysis for an experiment."""
    submission_file = exp.get("submission_file")
    if not submission_file or not Path(submission_file).exists():
        return None
    
    if not gold_by_transcript:
        return None
    
    pred_by_transcript = {}
    with open(submission_file) as f:
        for line in f:
            record = json.loads(line)
            tid = str(record['id'])
            pred_by_transcript[tid] = record.get('observations', [])
    
    stats = {"correct": 0, "wrong_value": 0, "missed": 0, "hallucinated": 0}
    
    for tid, gold_obs in gold_by_transcript.items():
        preds = pred_by_transcript.get(tid, [])
        valid_preds = [p for p in preds if isinstance(p, dict) and 'id' in p and 'value' in p]
        
        gold_by_concept = {str(g['id']): g for g in gold_obs}
        pred_by_concept = {str(p['id']): p for p in valid_preds}
        
        for gold in gold_obs:
            gid = str(gold['id'])
            if gid not in pred_by_concept:
                stats["missed"] += 1
            elif normalize_value(pred_by_concept[gid].get('value')) == normalize_value(gold.get('value')):
                stats["correct"] += 1
            else:
                stats["wrong_value"] += 1
        
        for pred in valid_preds:
            if str(pred['id']) not in gold_by_concept:
                stats["hallucinated"] += 1
    
    return stats


def deduplicate_experiments(experiments):
    """Remove duplicate experiments, keeping best F1 for each config."""
    unique = {}
    
    for exp in experiments:
        # Create unique key
        key = (
            exp["data_split"],
            exp["full_name"],
        )
        
        f1 = exp.get("f1") or 0
        if key not in unique or f1 > (unique[key].get("f1") or 0):
            unique[key] = exp
    
    return list(unique.values())


def filter_experiments(experiments, data_filter=None, model_filter=None):

    filtered = experiments
    
    if data_filter:
        filtered = [e for e in filtered if e["data_split"] == data_filter]
    
    if model_filter:
        model_lower = model_filter.lower()
        filtered = [e for e in filtered 
                   if model_lower in e.get("model", "").lower() 
                   or model_lower in e.get("full_name", "").lower()]
    
    return filtered


def export_csv(experiments, output_path):

    experiments = sorted(experiments, key=lambda x: (x.get("data_split") != "dev", -(x.get("f1") or 0)))
    
    # Compute error analysis
    gold_data_cache = {}
    for exp in experiments:
        split = exp.get("data_split", "unknown")
        if split not in gold_data_cache:
            gold_data_cache[split] = load_reference_data(split)
        exp["error_stats"] = compute_error_analysis(exp, gold_data_cache.get(split, {}))
    
    fieldnames = [
        "full_name", "data_split", "source", "precision", "recall", "f1", 
        "num_transcripts", "correct", "wrong_value", "missed", "hallucinated"
    ]
    
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for exp in experiments:
            row = {k: exp.get(k) for k in fieldnames}
            stats = exp.get("error_stats") or {}
            row["correct"] = stats.get("correct", "")
            row["wrong_value"] = stats.get("wrong_value", "")
            row["missed"] = stats.get("missed", "")
            row["hallucinated"] = stats.get("hallucinated", "")
            # Format metrics as percentages
            for metric in ["precision", "recall", "f1"]:
                if row.get(metric):
                    row[metric] = f"{row[metric]*100:.2f}%"
            writer.writerow(row)
    
    print(f"Saved {len(experiments)} experiments to {output_path}")


def export_excel(experiments, output_path):
    """Export experiments to Excel with separate sheets for dev/train."""
    if not HAS_OPENPYXL:
        print("Warning: openpyxl not installed, skipping Excel export")
        return
    
    # Separate by data split
    dev_exps = [e for e in experiments if e["data_split"] == "dev"]
    train_exps = [e for e in experiments if e["data_split"] in ["train", "unknown"]]
    
    # Sort by F1
    dev_exps = sorted(dev_exps, key=lambda x: -(x.get("f1") or 0))
    train_exps = sorted(train_exps, key=lambda x: -(x.get("f1") or 0))
    
    wb = openpyxl.Workbook()
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alt_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )
    
    def write_sheet(ws, exps, title):
        ws.title = title
        
        if not exps:
            ws.cell(row=1, column=1, value="No experiments found")
            return
        
        # Compute error analysis
        gold_data = load_reference_data("dev" if "Dev" in title else "train")
        for exp in exps:
            exp["error_stats"] = compute_error_analysis(exp, gold_data)
        
        # Headers
        headers = [
            ("Rank", 6),
            ("Method", 50),
            ("Type", 12),
            ("Precision", 11),
            ("Recall", 11),
            ("F1 Score", 11),
            ("Correct", 9),
            ("Wrong", 9),
            ("Missed", 9),
            ("Halluc.", 9),
        ]
        
        for col, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.freeze_panes = 'A2'
        
        best_f1 = max(e.get("f1") or 0 for e in exps) if exps else 0
        
        for row_idx, exp in enumerate(exps, 2):
            f1 = exp.get("f1") or 0
            is_best = abs(f1 - best_f1) < 0.0001
            stats = exp.get("error_stats") or {}
            
            values = [
                row_idx - 1,
                exp.get("full_name", ""),
                exp.get("source", "extraction"),
                f"{exp.get('precision', 0)*100:.2f}%" if exp.get("precision") else "N/A",
                f"{exp.get('recall', 0)*100:.2f}%" if exp.get("recall") else "N/A",
                f"{f1*100:.2f}%" if f1 else "N/A",
                stats.get("correct", ""),
                stats.get("wrong_value", ""),
                stats.get("missed", ""),
                stats.get("hallucinated", ""),
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                
                if col in [4, 5, 6, 7, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="right")
                elif col == 1:
                    cell.alignment = Alignment(horizontal="center")
                
                if is_best:
                    cell.fill = best_fill
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill
    
    # Create sheets
    ws_dev = wb.active
    write_sheet(ws_dev, dev_exps, "Dev Results")
    
    ws_train = wb.create_sheet()
    write_sheet(ws_train, train_exps, "Train Results")
    
    wb.save(output_path)
    print(f"Saved Excel report to {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Compare experiments")
    parser.add_argument("--data", choices=["train", "dev"], help="Filter by data split")
    parser.add_argument("--model", type=str, help="Filter by model name")
    parser.add_argument("--csv", type=str, help="Output CSV path")
    parser.add_argument("--excel", type=str, help="Output Excel path")
    parser.add_argument("--no-post-processing", action="store_true", 
                       help="Exclude voting/verification results")
    
    args = parser.parse_args()
    
    experiments = discover_experiments()
    print(f"Found {len(experiments)} extraction experiments")
    
    # Add post-processing results
    if not args.no_post_processing:
        pp_results = discover_post_processing_results()
        print(f"Found {len(pp_results)} post-processing results")
        experiments.extend(pp_results)
    
    # Deduplicate
    experiments = deduplicate_experiments(experiments)
    print(f"After deduplication: {len(experiments)} unique experiments")
    
    # Filter
    experiments = filter_experiments(experiments, args.data, args.model)
    if args.data or args.model:
        print(f"After filtering: {len(experiments)} experiments")
    
    csv_path = args.csv or (OUTPUTS_DIR / "comparison_results.csv")
    export_csv(experiments, csv_path)
    
    excel_path = args.excel or (OUTPUTS_DIR / "experiment_comparison.xlsx")
    export_excel(experiments, excel_path)


if __name__ == "__main__":
    main()
